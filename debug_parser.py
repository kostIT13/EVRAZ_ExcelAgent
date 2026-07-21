import openpyxl
import pandas as pd
import re

wb = openpyxl.load_workbook('data/data_proportional_prices.xlsx', data_only=True)
ws = wb['цв.лом(на окт25)']

# Читаем все данные
raw_data = []
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=False):
    raw_data.append([cell.value for cell in row])

# Разворачиваем merged cells
merged_map = {}
for merged_range in ws.merged_cells.ranges:
    min_row, min_col = merged_range.min_row - 1, merged_range.min_col - 1
    max_row, max_col = merged_range.max_row - 1, merged_range.max_col - 1
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            if (row, col) != (min_row, min_col):
                merged_map[(row, col)] = (min_row, min_col)

for row_idx, row in enumerate(raw_data):
    for col_idx in range(len(row)):
        if (row_idx, col_idx) in merged_map:
            src_row, src_col = merged_map[(row_idx, col_idx)]
            raw_data[row_idx][col_idx] = raw_data[src_row][src_col]

# Заголовки и данные
headers_raw = raw_data[:3]
data_rows = raw_data[3:]

# Берём первое не-None значение для каждой колонки
n_cols = max(len(row) for row in headers_raw)
columns = []
for col_idx in range(n_cols):
    col_name = None
    for row in headers_raw:
        if col_idx < len(row) and row[col_idx] is not None:
            val = str(row[col_idx]).strip()
            if val:
                col_name = val
                break
    if col_name is None:
        col_name = f'col_{col_idx}'
    columns.append(col_name)

# Создаём DataFrame
df = pd.DataFrame(data_rows, columns=columns)

print("=== ДО dropna ===")
print(f"Shape: {df.shape}")
print(f"Колонка 'Наименование лома' не-None: {df['Наименование лома'].notna().sum()} из {len(df)}")

# Симулируем dropna как в parser.py
df_clean = df.dropna(how='all', axis=1).dropna(how='all', axis=0)

print("\n=== ПОСЛЕ dropna ===")
print(f"Shape: {df_clean.shape}")
print(f"Колонки: {list(df_clean.columns)}")
if 'Наименование лома' in df_clean.columns:
    print(f"Колонка 'Наименование лома' не-None: {df_clean['Наименование лома'].notna().sum()} из {len(df_clean)}")
    print(df_clean['Наименование лома'].head(5))
