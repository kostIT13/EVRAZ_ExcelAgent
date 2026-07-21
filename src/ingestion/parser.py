"""
Парсинг Excel-файлов с обработкой merged cells и multi-level headers.
"""
import re
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd


def parse_excel(file_path: str | Path) -> dict[str, pd.DataFrame]:
    """
    Парсит Excel-файл и возвращает словарь {имя_листа: DataFrame}.
    
    Особенности:
    - Разворачивает merged cells (значение копируется во все ячейки диапазона)
    - Определяет multi-level headers (3 строки)
    - Преобразует заголовки в плоскую структуру
    - Извлекает месяц из названия листа
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    result = {}
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        df = _parse_sheet(ws, sheet_name)
        result[sheet_name] = df
    
    return result


def _parse_sheet(ws: openpyxl.worksheet, sheet_name: str) -> pd.DataFrame:
    """Парсит один лист Excel."""
    # 1. Читаем все данные в сыром виде
    raw_data = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=False):
        raw_data.append([cell.value for cell in row])
    
    # 2. Разворачиваем merged cells
    merged_map = _build_merged_map(ws)
    for row_idx, row in enumerate(raw_data):
        for col_idx in range(len(row)):
            if (row_idx, col_idx) in merged_map:
                src_row, src_col = merged_map[(row_idx, col_idx)]
                raw_data[row_idx][col_idx] = raw_data[src_row][src_col]
    
    # 3. Определяем заголовки (первые 3 строки)
    headers_raw = raw_data[:3]
    data_rows = raw_data[3:]
    
    # 4. Строим плоские имена колонок
    columns = _build_column_names(headers_raw)
    
    # 5. Создаём DataFrame
    df = pd.DataFrame(data_rows, columns=columns)
    
    # 6. Добавляем колонку с месяцем из названия листа
    month = _extract_month(sheet_name)
    df.insert(0, 'месяц', month)
    
    # 7. Удаляем пустые колонки и строки
    df = df.dropna(how='all', axis=1).dropna(how='all', axis=0)
    
    # 8. Конвертируем числовые колонки
    for col in df.columns:
        if col not in ('месяц', 'Наименование лома', '№ п/п'):
            df[col] = pd.to_numeric(df[col], errors='coerce')
    
    return df


def _build_merged_map(ws: openpyxl.worksheet) -> dict[tuple[int, int], tuple[int, int]]:
    """
    Строит карту merged cells.
    Для каждой ячейки, которая входит в объединённый диапазон,
    указывает координаты ячейки-источника (верхняя левая).
    """
    merged_map = {}
    for merged_range in ws.merged_cells.ranges:
        min_row, min_col = merged_range.min_row - 1, merged_range.min_col - 1
        max_row, max_col = merged_range.max_row - 1, merged_range.max_col - 1
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                if (row, col) != (min_row, min_col):
                    merged_map[(row, col)] = (min_row, min_col)
    return merged_map


def _build_column_names(headers_raw: list[list[Any]]) -> list[str]:
    """
    Строит плоские имена колонок из multi-level headers.
    Берёт первое не-None значение сверху вниз для каждой колонки.
    """
    n_cols = max(len(row) for row in headers_raw)
    columns = []
    
    for col_idx in range(n_cols):
        # Берём ПЕРВОЕ не-None значение сверху вниз
        col_name = None
        for row in headers_raw:
            if col_idx < len(row) and row[col_idx] is not None:
                val = str(row[col_idx]).strip()
                if val:
                    col_name = val
                    break  # берём первое и выходим
        
        if col_name is None:
            col_name = f'col_{col_idx}'
        
        # Если имя слишком длинное — проверяем, может это колонка аукциона
        if len(col_name) > 20:
            # Ищем уточнение в нижних строках
            for row in headers_raw[1:]:
                if col_idx < len(row) and row[col_idx] is not None:
                    val = str(row[col_idx]).strip()
                    if 'стартовая' in val:
                        col_name = 'аукцион_старт'
                        break
                    elif 'победителя' in val:
                        col_name = 'аукцион_победитель'
                        break
        
        columns.append(col_name)
    
    return columns




def _extract_month(sheet_name: str) -> str:
    """Извлекает месяц из названия листа, например 'цв.лом(на окт25)' → 'окт25'."""
    match = re.search(r'\(на\s+(\w+\d+)\)', sheet_name)
    if match:
        return match.group(1)
    return sheet_name


# Для тестирования
if __name__ == '__main__':
    dfs = parse_excel('data/data_proportional_prices.xlsx')
    for name, df in dfs.items():
        print(f'\n=== {name} ===')
        print(f'Shape: {df.shape}')
        print(f'Columns: {list(df.columns)}')
        print(df.head(3).to_string())
