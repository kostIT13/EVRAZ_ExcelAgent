from __future__ import annotations
import hashlib
import re
from pathlib import Path
from typing import Any, Dict, List, Optional
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from src.core.excel.schemas import ParsedCell, ParsedFile, ParsedHeader, ParsedSheet
from src.core.logging_settings import logger


_PHONE_PATTERN = re.compile(
    r'[\s]*[\(]?[\s]*\d{1,4}[\s\)\-]*\d{1,4}[\s\)\-]*\d{1,4}[\s\)\-]*\d{2,4}[\s\)\-]*\d{2,4}[\s\)\-]*',
)
_PAREN_PHONE_PATTERN = re.compile(r'\([^)]*\d{3,}[^)]*\)')
_PAREN_NAME_PATTERN = re.compile(r'\(\s*[А-Яа-яЁё]+\s*\)')
_STAR_PATTERN = re.compile(r'\*+')
_EXTRA_SPACE = re.compile(r'\s{2,}')


class ExcelParser:
    def __init__(self, file_path: Path, header_rows: Optional[int] = None):
        self.file_path = file_path
        self.workbook = None
        self._header_rows = header_rows

    def parse(self) -> ParsedFile:
        self.workbook = load_workbook(self.file_path, data_only=True)
        sheets = []

        for sheet_name in self.workbook.sheetnames:
            ws = self.workbook[sheet_name]
            logger.info("Parsing sheet: {}", sheet_name)

            self._unmerge_cells(ws)

            header_rows = self._header_rows or self._detect_header_rows(ws)
            logger.debug("Sheet '{}': detected {} header rows", sheet_name, header_rows)

            headers = self._parse_headers(ws, header_rows)

            headers = self._filter_empty_columns(ws, headers, header_rows)

            data = self._parse_data(ws, headers, header_rows)

            cells = self._collect_cells(ws, headers, header_rows)

            sheets.append(ParsedSheet(
                sheet_name=sheet_name,
                sheet_index=self.workbook.sheetnames.index(sheet_name),
                headers=headers,
                data=data,
                cells=cells,
                raw_data=self._get_raw_data(ws),
                header_rows=header_rows,
            ))

        total_rows = sum(s.row_count for s in sheets)
        total_cells = sum(s.col_count * s.row_count for s in sheets)

        return ParsedFile(
            filename=self.file_path.name,
            file_hash=self._get_file_hash(),
            sheets=sheets,
            total_rows=total_rows,
            total_cells=total_cells,
        )

    def _unmerge_cells(self, ws: Worksheet) -> None:
        merged = list(ws.merged_cells.ranges)
        if not merged:
            return

        logger.debug("Sheet '{}': unmerging {} ranges", ws.title, len(merged))

        ranges_with_values = []
        for merged_range in merged:
            top_left = (merged_range.min_row, merged_range.min_col)
            value = ws.cell(row=top_left[0], column=top_left[1]).value
            ranges_with_values.append((merged_range, value))

        for merged_range in merged:
            ws.unmerge_cells(str(merged_range))

        for merged_range, value in ranges_with_values:
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    ws.cell(row=row, column=col).value = value

        logger.debug("Sheet '{}': unmerged {} cells", ws.title, len(merged))

    def _detect_header_rows(self, ws: Worksheet) -> int:
        """Определяет число строк шапки.

        Сначала пробует эвристику (число в колонке A / плотность первых 5 колонок).
        Если она даёт шапку без осмысленных заголовков (например, лист начинается
        с блока заголовка-титула, как 'ШИХТОВОЧНЫЙ ЛИСТ'), то ищем настоящую строку
        заголовков по плотности заполнения ячеек по всей ширине листа.
        """
        max_col = ws.max_column or 1
        max_row = ws.max_row or 1

        default = self._default_header_rows(ws, max_col, max_row)

        # Если в найденной шапке есть хотя бы несколько осмысленных названий
        # колонок (не col_N) — это настоящая шапка, используем её как есть.
        if self._headers_are_meaningful(ws, default, max_col):
            return default

        dense = self._detect_dense_header_row(ws, max_col, max_row)
        if dense and dense != default:
            logger.debug(
                "Sheet header fallback: default={}, dense={}",
                default,
                dense,
            )
            return dense

        return default

    def _default_header_rows(self, ws: Worksheet, max_col: int, max_row: int) -> int:
        for row in range(1, min(max_row + 1, 10)):
            cell_val = ws.cell(row=row, column=1).value
            if isinstance(cell_val, (int, float)):
                return row - 1

        for row in range(1, min(max_row + 1, 10)):
            filled = 0
            for col in range(1, min(max_col + 1, 5)):
                if ws.cell(row=row, column=col).value is not None:
                    filled += 1
            if filled >= 3:
                if row > 1:
                    return row - 1

        return 3

    def _headers_are_meaningful(self, ws: Worksheet, header_rows: int, max_col: int) -> bool:
        """True, если шапка содержит достаточно РАЗНЫХ осмысленных заголовков.

        Считаем уникальные текстовые метки (не пустые и не auto col_N). Порог 4 —
        это отличает настоящую шапку с несколькими колонками (цены: наименование,
        поставщики, цена, аукцион) от одного повторяющегося заголовка-титула
        ('ШИХТОВОЧНЫЙ ЛИСТ'), растянутого через объединённые ячейки.
        """
        if header_rows <= 0:
            return False

        distinct_labels: set = set()
        for col in range(1, max_col + 1):
            levels = []
            for row in range(1, header_rows + 1):
                v = ws.cell(row=row, column=col).value
                if v is None:
                    continue
                s = str(v).strip().replace("\n", " ")
                if s and (not levels or s != levels[-1]):
                    levels.append(s)
            if not levels:
                continue
            full = self._clean_header_name(" > ".join(levels))
            if re.search(r"[а-яА-Яa-zA-Z]", full) and not re.fullmatch(r"col_\d+", full):
                distinct_labels.add(full.lower())
        return len(distinct_labels) >= 4

    def _detect_dense_header_row(self, ws: Worksheet, max_col: int, max_row: int) -> Optional[int]:
        """Ищет первую строку, плотно заполненную по всей ширине листа
        (признак настоящей шапки после блока титула). Возвращает номер этой строки."""
        limit = min(max_row, 30)
        if limit < 1:
            return None

        threshold = max(3, int(max_col * 0.2))

        for row in range(1, limit + 1):
            filled = 0
            for col in range(1, max_col + 1):
                v = ws.cell(row=row, column=col).value
                if v is not None and str(v).strip() != "":
                    filled += 1
            if filled >= threshold:
                return row

        return None

    def _filter_empty_columns(
        self,
        ws: Worksheet,
        headers: List[ParsedHeader],
        header_rows: int,
    ) -> List[ParsedHeader]:
        if not headers:
            return headers

        start_row = header_rows + 1
        max_row = ws.max_row or 0

        non_empty_headers: List[ParsedHeader] = []
        for header in headers:
            col = header.col_index
            has_data = False
            for row in range(start_row, max_row + 1):
                if ws.cell(row=row, column=col).value is not None:
                    has_data = True
                    break
            if has_data:
                non_empty_headers.append(header)
            else:
                logger.debug(
                    "Removing empty column {} ({})", col, header.full_name
                )

        return non_empty_headers

    def _get_raw_data(self, ws: Worksheet) -> List[List[Any]]:
        return [list(row) for row in ws.iter_rows(values_only=True)]

    def _parse_headers(self, ws: Worksheet, header_rows: int) -> List[ParsedHeader]:
        headers = []
        max_col = ws.max_column or 1

        for col in range(1, max_col + 1):
            raw_levels = []
            for row in range(1, header_rows + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value is not None:
                    cell_value = str(cell_value).strip().replace('\n', ' ')
                    raw_levels.append(cell_value)
                else:
                    raw_levels.append("")

            levels = []
            for lvl in raw_levels:
                if not levels or lvl != levels[-1]:
                    levels.append(lvl)

            full_name = " > ".join(l for l in levels if l) if any(l for l in levels if l) else f"col_{col}"

            full_name = self._clean_header_name(full_name)

            col_name = self._normalize_column_name(full_name)

            headers.append(ParsedHeader(
                levels=levels,
                full_name=full_name,
                col_index=col,
                col_name=col_name,
            ))

        return headers

    def _clean_header_name(self, name: str) -> str:
        name = _PAREN_PHONE_PATTERN.sub('', name)
        name = _PHONE_PATTERN.sub('', name)
        name = _PAREN_NAME_PATTERN.sub('', name)
        name = _STAR_PATTERN.sub('', name)
        name = _EXTRA_SPACE.sub(' ', name)
        name = name.strip(' ,;:.-*()')
        return name

    def _parse_data(self, ws: Worksheet, headers: List[ParsedHeader], header_rows: int) -> List[Dict[str, Any]]:
        data = []
        start_row = header_rows + 1
        max_row = ws.max_row or 0

        for row in range(start_row, max_row + 1):
            row_data: Dict[str, Any] = {}
            has_any_value = False

            for col_idx, header in enumerate(headers):
                col = header.col_index 
                cell_value = ws.cell(row=row, column=col).value

                if cell_value is not None:
                    has_any_value = True
                    if isinstance(cell_value, (int, float)):
                        row_data[header.col_name] = cell_value
                    else:
                        row_data[header.col_name] = str(cell_value).strip()
                else:
                    row_data[header.col_name] = None

            if has_any_value:
                data.append(row_data)

        return data

    def _collect_cells(self, ws: Worksheet, headers: List[ParsedHeader], header_rows: int) -> List[ParsedCell]:
        cells = []
        start_row = header_rows + 1
        max_row = ws.max_row or 0

        for row in range(start_row, max_row + 1):
            for header in headers:
                col = header.col_index
                cell_value = ws.cell(row=row, column=col).value
                if cell_value is not None:
                    cells.append(ParsedCell(
                        row=row,
                        col=col,
                        value=cell_value,
                        col_name=header.col_name,
                        sheet_name=ws.title,
                    ))

        return cells

    def _normalize_column_name(self, name: str) -> str:
        name = re.sub(r'[^\w\s]', '', name)
        name = re.sub(r'\s+', '_', name)
        name = name.lower()
        name = re.sub(r'_+', '_', name)
        name = name.strip('_')
        return name or "unknown"

    def _get_file_hash(self) -> str:
        with open(self.file_path, 'rb') as f:
            return hashlib.sha256(f.read()).hexdigest()[:16]