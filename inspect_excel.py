import openpyxl

wb = openpyxl.load_workbook('data/data_proportional_prices.xlsx', data_only=True)

print('=== СПИСОК ЛИСТОВ ===')
print(wb.sheetnames)

for name in wb.sheetnames:
    ws = wb[name]
    print(f'\n{"="*60}')
    print(f'=== ЛИСТ: {name} ===')
    print(f'Размер: {ws.dimensions}')
    print(f'Строк: {ws.max_row}, Колонок: {ws.max_column}')
    
    # Покажем первые 20 строк
    print(f'\n--- Первые 20 строк ---')
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(20, ws.max_row), values_only=False), 1):
        vals = []
        for cell in row:
            v = cell.value
            if v is not None:
                vals.append(f'{cell.coordinate}={repr(v)}')
        if vals:
            print(f'  Row {row_idx}: {vals}')
    
    # Проверим merged cells
    if ws.merged_cells.ranges:
        print(f'\n--- Объединённые ячейки ({len(ws.merged_cells.ranges)} шт) ---')
        for m in ws.merged_cells.ranges:
            print(f'  {m}')
    else:
        print('\n--- Объединённых ячеек нет ---')
